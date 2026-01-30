"""
Historical Odds Builder - Tennis Betting System

Downloads tennis-data.co.uk XLSX files (ATP + WTA main tour),
matches to database matches using PlayerNameMatcher, and outputs
a JSON lookup file for use by the cloud backtester.

Usage:
    python odds_builder.py --db-path /path/to/tennis_betting.db
    python odds_builder.py --db-path /path/to/db --output odds_lookup.json
    python odds_builder.py --db-path /path/to/db --years 2024 2025
"""

import sys
import json
import sqlite3
import argparse
import tempfile
import urllib.request
from pathlib import Path
from datetime import datetime, timedelta
from typing import Optional, Dict, List

sys.stdout.reconfigure(encoding='utf-8', errors='replace')


# =============================================================================
# PlayerNameMatcher (embedded from tennis_explorer_scraper.py lines 14-351)
# Self-contained, only depends on typing.Optional
# =============================================================================

class PlayerNameMatcher:
    """
    Robust player name matching system that handles various name formats:
    - "LastName F." (e.g., "Grubor A.")
    - "F. LastName" (e.g., "A. Grubor")
    - "FirstName LastName" (e.g., "Ana Grubor")
    - Compound names (e.g., "Del Potro J.", "Juan Martin Del Potro")
    """

    def __init__(self):
        self.players = {}
        self.player_rankings = {}
        self.by_full_name = {}
        self.by_last_name = {}
        self.by_name_parts = {}
        self.by_last_initial = {}

    def _normalize(self, name: str) -> str:
        if not name:
            return ""
        name = name.lower().strip()
        name = name.replace('.', '')
        name = ' '.join(name.split())
        return name

    def _extract_components(self, name: str) -> dict:
        result = {
            'last_name': '', 'first_name': '',
            'first_initial': '', 'all_parts': []
        }
        if not name:
            return result
        normalized = self._normalize(name)
        parts = normalized.split()
        result['all_parts'] = parts
        if not parts:
            return result
        if len(parts) == 1:
            result['last_name'] = parts[0]
            return result
        first_is_initial = len(parts[0]) == 1
        last_is_initial = len(parts[-1]) == 1
        if last_is_initial:
            result['first_initial'] = parts[-1]
            result['last_name'] = parts[0]
            if len(parts) > 2:
                result['first_name'] = ' '.join(parts[1:-1])
        elif first_is_initial:
            result['first_initial'] = parts[0]
            result['last_name'] = parts[-1]
            if len(parts) > 2:
                result['first_name'] = ' '.join(parts[1:-1])
        else:
            result['first_name'] = parts[0]
            result['last_name'] = parts[-1]
            result['first_initial'] = parts[0][0] if parts[0] else ''
        return result

    def load_players(self, db_connection):
        cursor = db_connection.cursor()
        cursor.execute("SELECT id, name, current_ranking FROM players")
        for row in cursor.fetchall():
            self.add_player(row[0], row[1], row[2])

    def add_player(self, player_id: int, full_name: str, current_ranking: int = None):
        if not full_name:
            return
        self.players[player_id] = full_name
        self.player_rankings[player_id] = current_ranking
        normalized = self._normalize(full_name)
        components = self._extract_components(full_name)
        self.by_full_name[normalized] = player_id
        self.by_full_name[normalized.replace(' ', '')] = player_id
        for part in components['all_parts']:
            if len(part) > 1:
                if part not in self.by_name_parts:
                    self.by_name_parts[part] = []
                self.by_name_parts[part].append((player_id, full_name))
        last_name = components['last_name']
        if last_name and len(last_name) > 1:
            if last_name not in self.by_last_name:
                self.by_last_name[last_name] = []
            first_initial = components['first_initial'] or (
                components['first_name'][0] if components['first_name'] else '')
            self.by_last_name[last_name].append((player_id, full_name, first_initial))
            if first_initial:
                key = f"{last_name}_{first_initial}"
                if key not in self.by_last_initial:
                    self.by_last_initial[key] = []
                self.by_last_initial[key].append((player_id, full_name))
        for part in components['all_parts']:
            if len(part) > 1:
                if part not in self.by_last_name:
                    self.by_last_name[part] = []
                all_initials = set()
                for p in components['all_parts']:
                    if len(p) == 1:
                        all_initials.add(p)
                    elif p != part and len(p) > 1:
                        all_initials.add(p[0])
                existing = [(pid, fn) for pid, fn, fi in self.by_last_name[part]]
                if (player_id, full_name) not in existing:
                    first_initial = list(all_initials)[0] if all_initials else ''
                    self.by_last_name[part].append((player_id, full_name, first_initial))
                for initial in all_initials:
                    key = f"{part}_{initial}"
                    if key not in self.by_last_initial:
                        self.by_last_initial[key] = []
                    if (player_id, full_name) not in self.by_last_initial[key]:
                        self.by_last_initial[key].append((player_id, full_name))

    def find_player_id(self, name: str) -> Optional[int]:
        if not name:
            return None
        normalized = self._normalize(name)
        components = self._extract_components(name)
        if normalized in self.by_full_name:
            exact_match_id = self.by_full_name[normalized]
            if self.player_rankings.get(exact_match_id) is not None:
                return exact_match_id
        no_spaces = normalized.replace(' ', '')
        if no_spaces in self.by_full_name:
            exact_match_id = self.by_full_name[no_spaces]
            if self.player_rankings.get(exact_match_id) is not None:
                return exact_match_id
        significant_parts = sorted(
            [p for p in components['all_parts'] if len(p) > 1],
            key=len, reverse=True)
        initial = None
        for p in components['all_parts']:
            if len(p) == 1:
                initial = p
                break
        if not initial and len(significant_parts) >= 2:
            initial = min(significant_parts, key=len)[0]
        for part in significant_parts:
            if len(part) < 3:
                continue
            if part in self.by_last_name:
                candidates = self.by_last_name[part]
                if initial:
                    matching = [(pid, fn) for pid, fn, fi in candidates
                               if fi and fi[0] == initial]
                    if matching:
                        return self._pick_best_candidate(matching)
                if len(candidates) == 1:
                    return candidates[0][0]
        last_name = components['last_name']
        first_initial = components['first_initial'] or initial
        if last_name and len(last_name) >= 2 and first_initial:
            key = f"{last_name}_{first_initial}"
            if key in self.by_last_initial:
                candidates = self.by_last_initial[key]
                return self._pick_best_candidate(candidates)
        if len(significant_parts) >= 2:
            long_parts = [p for p in significant_parts if len(p) >= 3]
            if len(long_parts) >= 2:
                matching_players = []
                for pid, full_name in self.players.items():
                    fn_normalized = self._normalize(full_name)
                    fn_parts = fn_normalized.split()
                    matches = sum(1 for sp in long_parts
                                 if any(sp == fp or sp in fp or fp in sp for fp in fn_parts))
                    if matches == len(long_parts):
                        matching_players.append((pid, full_name))
                if matching_players:
                    return self._pick_best_candidate(matching_players)
        if len(significant_parts) == 1 and initial:
            part = significant_parts[0]
            if len(part) >= 3 and part in self.by_last_name:
                candidates = self.by_last_name[part]
                matching = [(pid, fn) for pid, fn, fi in candidates
                           if fi and fi[0] == initial]
                if matching:
                    return self._pick_best_candidate(matching)
        if len(significant_parts) >= 2:
            reversed_name = ' '.join(significant_parts[::-1])
            if reversed_name in self.by_full_name:
                return self.by_full_name[reversed_name]
            swapped = f"{significant_parts[-1]} {' '.join(significant_parts[:-1])}"
            if swapped in self.by_full_name:
                return self.by_full_name[swapped]
        if normalized in self.by_full_name:
            return self.by_full_name[normalized]
        if no_spaces in self.by_full_name:
            return self.by_full_name[no_spaces]
        return None

    def _pick_best_candidate(self, candidates: list) -> Optional[int]:
        if not candidates:
            return None
        if len(candidates) == 1:
            return candidates[0][0]
        player_ids = [c[0] for c in candidates]
        ranked_players = [(pid, self.player_rankings.get(pid)) for pid in player_ids
                         if self.player_rankings.get(pid) is not None]
        if ranked_players:
            ranked_players.sort(key=lambda x: x[1])
            return ranked_players[0][0]
        positive_ids = [pid for pid in player_ids if pid > 0]
        if positive_ids:
            return positive_ids[0]
        return player_ids[0]


# =============================================================================
# Odds Builder
# =============================================================================

def safe_float(val) -> Optional[float]:
    """Safely convert a value to float, returning None on failure."""
    if val is None:
        return None
    try:
        f = float(val)
        return f if f > 1.0 else None
    except (ValueError, TypeError):
        return None


def download_file(url: str, filepath: Path) -> bool:
    """Download a file from URL. Returns True on success."""
    try:
        print(f"  Downloading {url}...")
        urllib.request.urlretrieve(url, str(filepath))
        return True
    except Exception as e:
        print(f"  Failed: {e}")
        return False


def download_files(years: List[int], download_dir: Path) -> List[Path]:
    """Download ATP and WTA XLSX files from tennis-data.co.uk."""
    files = []
    for year in years:
        # ATP
        atp_url = f"http://www.tennis-data.co.uk/{year}/{year}.xlsx"
        atp_path = download_dir / f"atp_{year}.xlsx"
        if download_file(atp_url, atp_path):
            files.append(atp_path)

        # WTA (available from 2007)
        if year >= 2007:
            wta_url = f"http://www.tennis-data.co.uk/{year}w/{year}.xlsx"
            wta_path = download_dir / f"wta_{year}.xlsx"
            if download_file(wta_url, wta_path):
                files.append(wta_path)

    return files


def parse_xlsx(filepath: Path) -> List[Dict]:
    """Parse a tennis-data XLSX file into standardized rows."""
    from openpyxl import load_workbook

    print(f"  Parsing {filepath.name}...")
    wb = load_workbook(str(filepath), read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return []

    headers = [str(h).strip() if h else '' for h in rows[0]]
    results = []

    for row in rows[1:]:
        if len(row) < len(headers):
            continue
        row_dict = dict(zip(headers, row))

        # Parse date
        date_val = row_dict.get('Date')
        date_str = None
        if isinstance(date_val, datetime):
            date_str = date_val.strftime('%Y-%m-%d')
        elif isinstance(date_val, str):
            for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%m/%d/%Y'):
                try:
                    date_str = datetime.strptime(date_val.strip(), fmt).strftime('%Y-%m-%d')
                    break
                except ValueError:
                    continue
        if not date_str:
            continue

        winner = str(row_dict.get('Winner', '')).strip()
        loser = str(row_dict.get('Loser', '')).strip()
        if not winner or not loser:
            continue

        # Extract odds: Pinnacle > Average > Bet365
        winner_odds = loser_odds = None
        source = None

        psw = safe_float(row_dict.get('PSW'))
        psl = safe_float(row_dict.get('PSL'))
        if psw and psl:
            winner_odds, loser_odds, source = psw, psl, "PIN"
        else:
            avgw = safe_float(row_dict.get('AvgW'))
            avgl = safe_float(row_dict.get('AvgL'))
            if avgw and avgl:
                winner_odds, loser_odds, source = avgw, avgl, "AVG"
            else:
                b365w = safe_float(row_dict.get('B365W'))
                b365l = safe_float(row_dict.get('B365L'))
                if b365w and b365l:
                    winner_odds, loser_odds, source = b365w, b365l, "B365"

        if not winner_odds:
            continue

        results.append({
            'date': date_str,
            'winner': winner,
            'loser': loser,
            'tournament': str(row_dict.get('Tournament', '')).strip(),
            'surface': str(row_dict.get('Surface', '')).strip(),
            'winner_odds': round(winner_odds, 2),
            'loser_odds': round(loser_odds, 2),
            'source': source,
        })

    wb.close()
    print(f"    {len(results)} matches with odds")
    return results


def build_match_index(conn) -> Dict:
    """Build lookup: (date, winner_id, loser_id) -> match_id."""
    cursor = conn.cursor()
    cursor.execute("SELECT id, date, winner_id, loser_id FROM matches")

    index = {}
    for match_id, date_str, winner_id, loser_id in cursor.fetchall():
        if not date_str or not winner_id or not loser_id:
            continue
        # Normalize date to YYYY-MM-DD
        if isinstance(date_str, str) and len(date_str) >= 10:
            date_str = date_str[:10]
        index[(date_str, winner_id, loser_id)] = match_id

    print(f"  Built match index: {len(index)} matches")
    return index


def match_rows_to_db(rows: List[Dict], matcher: PlayerNameMatcher,
                     match_index: Dict) -> Dict:
    """Match tennis-data.co.uk rows to database matches and extract odds."""
    odds = {}
    stats = {
        'total': len(rows),
        'matched': 0,
        'name_fail': 0,
        'match_fail': 0,
        'winner_fail': 0,
        'loser_fail': 0,
    }
    source_counts = {'PIN': 0, 'AVG': 0, 'B365': 0}
    unmatched_names = []

    for row in rows:
        winner_id = matcher.find_player_id(row['winner'])
        loser_id = matcher.find_player_id(row['loser'])

        if not winner_id:
            stats['winner_fail'] += 1
            stats['name_fail'] += 1
            if len(unmatched_names) < 30:
                unmatched_names.append(f"Winner: {row['winner']}")
            continue
        if not loser_id:
            stats['loser_fail'] += 1
            stats['name_fail'] += 1
            if len(unmatched_names) < 30:
                unmatched_names.append(f"Loser: {row['loser']}")
            continue

        # Look up match in database
        date_str = row['date']
        match_id = None

        # Try exact date
        key = (date_str, winner_id, loser_id)
        if key in match_index:
            match_id = match_index[key]
        else:
            # Try +/- 1 day (timezone/date boundary differences)
            try:
                dt = datetime.strptime(date_str, '%Y-%m-%d')
                for delta in [1, -1]:
                    alt_date = (dt + timedelta(days=delta)).strftime('%Y-%m-%d')
                    alt_key = (alt_date, winner_id, loser_id)
                    if alt_key in match_index:
                        match_id = match_index[alt_key]
                        break
            except ValueError:
                pass

        if not match_id:
            stats['match_fail'] += 1
            continue

        stats['matched'] += 1
        source_counts[row['source']] = source_counts.get(row['source'], 0) + 1

        odds[str(match_id)] = {
            'w': row['winner_odds'],
            'l': row['loser_odds'],
            'src': row['source'],
        }

    return {
        '_meta': {
            'generated': datetime.now().isoformat(),
            'total_rows': stats['total'],
            'matched': stats['matched'],
            'name_failures': stats['name_fail'],
            'match_failures': stats['match_fail'],
            'winner_name_failures': stats['winner_fail'],
            'loser_name_failures': stats['loser_fail'],
            'sources': source_counts,
            'match_rate': f"{stats['matched'] / stats['total'] * 100:.1f}%"
                         if stats['total'] > 0 else '0%',
        },
        '_unmatched_names': unmatched_names[:30],
        'odds': odds,
    }


def detect_years_from_db(db_path: str) -> List[int]:
    """Detect which years have matches in the database."""
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT DISTINCT SUBSTR(date, 1, 4) as year
        FROM matches
        WHERE date IS NOT NULL AND LENGTH(date) >= 4
        ORDER BY year
    """)
    years = [int(row[0]) for row in cursor.fetchall() if row[0].isdigit()]
    conn.close()
    return years


def main():
    parser = argparse.ArgumentParser(description='Build historical odds lookup from tennis-data.co.uk')
    parser.add_argument('--db-path', type=str, required=True,
                        help='Path to tennis_betting.db')
    parser.add_argument('--output', type=str, default='odds_lookup.json',
                        help='Output JSON file path (default: odds_lookup.json)')
    parser.add_argument('--years', type=int, nargs='+', default=None,
                        help='Years to download (default: auto-detect from DB)')
    parser.add_argument('--download-dir', type=str, default=None,
                        help='Directory for downloaded files (default: temp dir)')
    parser.add_argument('--skip-download', action='store_true',
                        help='Skip download, use existing files in download-dir')
    args = parser.parse_args()

    print("=" * 60)
    print("  Historical Odds Builder")
    print("=" * 60)

    # Validate database
    db_path = Path(args.db_path)
    if not db_path.exists():
        print(f"ERROR: Database not found: {db_path}")
        sys.exit(1)

    # Determine years
    if args.years:
        years = args.years
    else:
        years = detect_years_from_db(str(db_path))
        # Filter to reasonable range (tennis-data.co.uk has data from 2000+)
        years = [y for y in years if 2000 <= y <= datetime.now().year]

    if not years:
        print("ERROR: No valid years found")
        sys.exit(1)

    print(f"\n  Years: {min(years)}-{max(years)} ({len(years)} years)")
    print(f"  Database: {db_path}")
    print(f"  Output: {args.output}")

    # Download XLSX files
    download_dir = Path(args.download_dir) if args.download_dir else Path(tempfile.mkdtemp())
    download_dir.mkdir(parents=True, exist_ok=True)

    if not args.skip_download:
        print(f"\n[1/4] Downloading XLSX files to {download_dir}...")
        xlsx_files = download_files(years, download_dir)
        print(f"  Downloaded {len(xlsx_files)} files")
    else:
        xlsx_files = list(download_dir.glob("*.xlsx"))
        print(f"\n[1/4] Using existing files: {len(xlsx_files)} found")

    # Parse all XLSX files
    print("\n[2/4] Parsing XLSX files...")
    all_rows = []
    for filepath in sorted(xlsx_files):
        all_rows.extend(parse_xlsx(filepath))
    print(f"  Total: {len(all_rows)} matches with odds data")

    if not all_rows:
        print("ERROR: No odds data found")
        sys.exit(1)

    # Build matcher and match index from database
    print("\n[3/4] Building player matcher and match index...")
    conn = sqlite3.connect(str(db_path))
    matcher = PlayerNameMatcher()
    matcher.load_players(conn)
    print(f"  Loaded {len(matcher.players)} players into matcher")

    match_index = build_match_index(conn)
    conn.close()

    # Match rows to database
    print("\n[4/4] Matching odds to database matches...")
    result = match_rows_to_db(all_rows, matcher, match_index)

    # Write output
    output_path = Path(args.output)
    with open(output_path, 'w') as f:
        json.dump(result, f, separators=(',', ':'))

    file_size = output_path.stat().st_size
    print(f"\n  Written: {output_path} ({file_size / 1024:.0f} KB)")

    # Print statistics
    meta = result['_meta']
    print(f"\n{'=' * 60}")
    print("  RESULTS")
    print(f"{'=' * 60}")
    print(f"  Total odds rows:    {meta['total_rows']}")
    print(f"  Matched to DB:      {meta['matched']} ({meta['match_rate']})")
    print(f"  Name failures:      {meta['name_failures']} "
          f"(W: {meta['winner_name_failures']}, L: {meta['loser_name_failures']})")
    print(f"  Match failures:     {meta['match_failures']} (players found, match not in DB)")
    print(f"\n  Odds sources:")
    for src, count in sorted(meta['sources'].items()):
        print(f"    {src}: {count}")

    if result['_unmatched_names']:
        print(f"\n  Unmatched names (first {len(result['_unmatched_names'])}):")
        for name in result['_unmatched_names']:
            print(f"    {name}")

    print(f"\n{'=' * 60}")


if __name__ == '__main__':
    main()
